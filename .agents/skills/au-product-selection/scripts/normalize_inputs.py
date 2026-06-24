#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Normalize SellerSprite/Sorftime/TikTok/Trend CSV, XLSX, or JSON files for AU product selection."""
import argparse
import csv
import json
import re
import shutil
from datetime import date
from pathlib import Path

ALIASES = {
    "candidate": ["candidate", "product", "product_name", "\u54c1\u540d", "\u4ea7\u54c1", "\u4ea7\u54c1\u540d\u79f0", "\u5546\u54c1\u6807\u9898", "title", "\u6807\u9898"],
    "primary_keyword": ["primary_keyword", "keyword", "\u5173\u952e\u8bcd", "\u4e3b\u5173\u952e\u8bcd", "\u641c\u7d22\u8bcd"],
    "long_tail_keywords": ["long_tail_keywords", "related_keywords", "\u957f\u5c3e\u8bcd", "\u7ec6\u5206\u5173\u952e\u8bcd", "\u76f8\u5173\u5173\u952e\u8bcd", "AC\u5173\u952e\u8bcd"],
    "monthly_sales": ["monthly_sales", "\u6708\u9500\u91cf", "\u6708\u9500\u552e\u91cf", "top100_monthly_sales"],
    "monthly_sales_growth": ["monthly_sales_growth", "\u6708\u9500\u91cf\u589e\u957f\u7387"],
    "monthly_revenue": ["monthly_revenue", "\u6708\u9500\u552e\u989d", "\u6708\u9500\u552e\u989d($)", "\u9500\u552e\u989d"],
    "search_volume": ["search_volume", "monthly_search", "\u6708\u641c\u7d22\u91cf", "\u641c\u7d22\u91cf"],
    "top3_product_concentration": ["top3_product_concentration", "\u5546\u54c1\u96c6\u4e2d\u5ea6", "\u94fe\u63a5\u96c6\u4e2d\u5ea6"],
    "top3_brand_concentration": ["top3_brand_concentration", "\u54c1\u724c\u96c6\u4e2d\u5ea6"],
    "top3_seller_concentration": ["top3_seller_concentration", "\u5356\u5bb6\u96c6\u4e2d\u5ea6", "\u5e97\u94fa\u96c6\u4e2d\u5ea6"],
    "click_concentration": ["click_concentration", "\u70b9\u51fb\u96c6\u4e2d\u5ea6"],
    "conversion_concentration": ["conversion_concentration", "\u8f6c\u5316\u96c6\u4e2d\u5ea6", "\u8f6c\u5316\u603b\u5360\u6bd4"],
    "spr": ["spr", "SPR"],
    "cpc": ["cpc", "PPC", "CPC", "\u63a8\u8350CPC"],
    "review_count": ["review_count", "\u8bc4\u8bba\u6570", "\u8bc4\u5206\u6570"],
    "rating": ["rating", "\u8bc4\u5206", "\u661f\u7ea7"],
    "seller_count": ["seller_count", "\u5356\u5bb6\u6570"],
    "variant_count": ["variant_count", "\u53d8\u4f53\u6570"],
    "amazon_owned_share": ["amazon_owned_share", "Amazon\u81ea\u8425\u5360\u6bd4", "\u4e9a\u9a6c\u900a\u81ea\u8425\u5360\u6bd4"],
    "new_product_count": ["new_product_count", "\u65b0\u54c1\u6570\u91cf", "\u534a\u5e74\u65b0\u54c1\u6570"],
    "new_product_sales_share": ["new_product_sales_share", "\u65b0\u54c1\u9500\u91cf\u5360\u6bd4"],
    "price": ["price", "\u552e\u4ef7", "\u4ef7\u683c", "\u4ef7\u683c($)", "\u5ba2\u5355\u4ef7"],
    "fba_fee": ["fba_fee", "FBA($)", "FBA"],
    "cost": ["cost", "\u91c7\u8d2d\u6210\u672c", "\u6210\u672c"],
    "gross_margin": ["gross_margin", "\u5229\u6da6\u7387", "\u6bdb\u5229\u7387"],
    "return_rate": ["return_rate", "\u9000\u8d27\u7387"],
    "forbidden_flags": ["forbidden_flags", "\u7981\u5165", "\u98ce\u9669\u6807\u7b7e", "\u6807\u7b7e"],
    "patent_risk": ["patent_risk", "\u4e13\u5229\u98ce\u9669"],
    "compliance_risk": ["compliance_risk", "\u5408\u89c4\u98ce\u9669"],
    "asin": ["ASIN", "asin"],
    "brand": ["\u54c1\u724c", "brand"],
    "product_url": ["\u5546\u54c1\u8be6\u60c5\u9875\u94fe\u63a5", "product_url", "url"],
    "image_url": ["\u5546\u54c1\u4e3b\u56fe", "image_url"],
    "category_path": ["\u7c7b\u76ee\u8def\u5f84", "category_path"],
    "main_category": ["\u5927\u7c7b\u76ee", "main_category"],
    "sub_category": ["\u5c0f\u7c7b\u76ee", "sub_category"],
    "launch_days": ["\u4e0a\u67b6\u5929\u6570", "launch_days"],
    "package_size_segment": ["\u5305\u88c5\u5c3a\u5bf8\u5206\u6bb5", "package_size_segment"],
    "detail_params": ["\u8be6\u7ec6\u53c2\u6570", "detail_params"],
}
MONTH_RE = re.compile(r"^(20\d{2})[-/\u5e74.]?(0?[1-9]|1[0-2])\u6708?$")
WORD_RE = re.compile(r"[A-Za-z][A-Za-z0-9+\-']*")
STOPWORDS = {"for", "and", "with", "the", "a", "an", "of", "to", "in", "on", "by", "pack", "set", "pcs", "count", "new", "upgraded", "compatible"}
RISK_KEYWORDS = {
    "sunscreen": "topical/liquid-compliance risk", "spf": "topical/liquid-compliance risk", "cream": "topical/liquid-compliance risk",
    "oil": "liquid/topical risk", "spray": "liquid/aerosol risk", "serum": "liquid/topical risk", "cleanser": "topical/health claim risk",
    "soap": "topical/health claim risk", "medicated": "medical claim risk", "antifungal": "medical claim risk", "treatment": "medical claim risk",
    "vitamin": "supplement risk", "supplement": "supplement risk", "capsule": "supplement risk", "gummy": "supplement/food risk",
    "drops": "liquid risk", "food": "food/import risk", "baby": "child-safety risk", "kids": "child-safety risk",
}
PRODUCT_NOUNS = {
    "speaker", "machine", "stick", "shelf", "rack", "case", "holder", "organizer", "cleanser", "soap", "heads", "brush",
    "display", "fan", "trap", "charger", "splitter", "repeller", "system", "refill", "grinder",
}


def read_csv(path):
    for enc in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                return list(csv.DictReader(f))
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("csv", b"", 0, 1, f"cannot decode {path}")


def read_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read .xlsx inputs") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = [ws for ws in workbook.worksheets if ws.title.strip().lower() != "notes"]
    if not sheets:
        return []
    preferred = [ws for ws in sheets if "product" in ws.title.lower()]
    ws = preferred[0] if preferred else sheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    headers = ["" if h is None else str(h).strip() for h in headers]
    records = []
    for row in rows:
        record = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else None
            record[header] = value
        if any(v not in (None, "") for v in record.values()):
            records.append(record)
    return records


def read_json(path):
    for enc in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return json.loads(path.read_text(encoding=enc))
        except UnicodeDecodeError:
            continue
    raise ValueError(f"cannot decode {path}")


def as_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value)
    text = raw.strip().replace(",", "").replace("$", "").replace("\uffe5", "")
    is_percent = "%" in text
    text = text.replace("%", "")
    try:
        num = float(text)
    except ValueError:
        return None
    if is_percent or (0 < num <= 100 and any(k in raw for k in ("%", "\u5360\u6bd4", "\u7387", "\u96c6\u4e2d\u5ea6"))):
        return num / 100
    return num


def pick(row, field):
    lower = {str(k).strip().lower(): v for k, v in row.items()}
    for alias in ALIASES[field]:
        if alias in row and row[alias] not in (None, ""):
            return row[alias]
        val = lower.get(alias.lower())
        if val not in (None, ""):
            return val
    return None


def split_keywords(value):
    if not value:
        return []
    if isinstance(value, list):
        return [str(x).strip() for x in value if str(x).strip()]
    return [x.strip() for x in re.split(r"[,;\uff0c\uff1b|/\n\r]+", str(value)) if x.strip()]


def unique_keep_order(items):
    seen = set()
    out = []
    for item in items:
        normalized = str(item).strip()
        key = normalized.lower()
        if normalized and key not in seen:
            seen.add(key)
            out.append(normalized)
    return out


def keyword_tokens(text):
    return [w for w in WORD_RE.findall(str(text or "")) if w.lower() not in STOPWORDS]


def strip_brand_from_keyword(keyword, brand=None):
    words = keyword_tokens(keyword)
    brand_words = [w.lower() for w in keyword_tokens(brand)]
    if brand_words and [w.lower() for w in words[:len(brand_words)]] == brand_words:
        words = words[len(brand_words):]
    return " ".join(words)


def is_valid_keyword_seed(keyword):
    return len(keyword_tokens(keyword)) >= 2


def clean_keyword_seed(keyword, brand=None):
    cleaned = strip_brand_from_keyword(keyword, brand)
    return cleaned if is_valid_keyword_seed(cleaned) else ""


def clean_keyword_list(keywords, brand=None):
    cleaned = [clean_keyword_seed(keyword, brand) for keyword in keywords]
    return unique_keep_order([keyword for keyword in cleaned if keyword])


def title_tokens_legacy(title):
    """Backward-compatible alias for older ad hoc imports."""
    return title_tokens(title)


def title_tokens(title, brand=None):
    words = keyword_tokens(title)
    brand_words = [w.lower() for w in keyword_tokens(brand)]
    if not brand_words:
        return words
    cleaned = []
    i = 0
    while i < len(words):
        window = [w.lower() for w in words[i:i + len(brand_words)]]
        if window == brand_words:
            i += len(brand_words)
            continue
        cleaned.append(words[i])
        i += 1
    return cleaned


def extract_title_seed_keywords(title, brand=None):
    words = title_tokens(title, brand)
    seeds = []
    if len(words) >= 4 and words[3].lower() in PRODUCT_NOUNS:
        seeds.append(" ".join(words[:4]))
    # First meaningful 2-3 word phrases usually contain the product noun phrase without crossing too far into attributes.
    for n in (3, 2):
        if len(words) >= n:
            seeds.append(" ".join(words[:n]))
    lower_words = [w.lower() for w in words]
    for i, word in enumerate(lower_words):
        if word in PRODUCT_NOUNS:
            start = max(0, i - 2)
            seeds.append(" ".join(words[start:i + 1]))
    # Pair important modifiers with product nouns, e.g. sunscreen stick / sun stick / display shelf.
    for left in ("sun", "sunscreen", "bottle", "liquor", "shot", "toothbrush", "replacement", "pillow", "white", "bluetooth"):
        for right in ("speaker", "machine", "stick", "shelf", "rack", "case", "heads", "brush", "holder", "display"):
            if any(lower_words[i:i + 2] == [left, right] for i in range(len(lower_words) - 1)):
                seeds.append(f"{left} {right}")
    return clean_keyword_list(seeds, brand)[:12]


def risk_flags_from_text(*parts):
    text = " ".join(str(p or "") for p in parts).lower()
    flags = []
    for term, label in RISK_KEYWORDS.items():
        if term in text:
            flags.append(f"{term}: {label}")
    return unique_keep_order(flags)


def month_key(header):
    m = MONTH_RE.match(str(header).strip())
    if not m:
        return None
    return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"


def extract_history(row, primary_keyword, source):
    history = []
    for key, value in row.items():
        mk = month_key(key)
        if not mk:
            continue
        num = as_number(value)
        if num is not None:
            history.append({"keyword": primary_keyword, "month": mk, "search_volume": int(num), "source": source})
    return sorted(history, key=lambda x: x["month"])


def normalize_row(row, source_file, row_number, source_site, target_site):
    title = pick(row, "candidate")
    brand = pick(row, "brand")
    ac_keywords = clean_keyword_list(split_keywords(pick(row, "long_tail_keywords")), brand)
    title_seeds = extract_title_seed_keywords(title, brand)
    picked_primary = clean_keyword_seed(pick(row, "primary_keyword"), brand)
    primary = picked_primary or (title_seeds[0] if title_seeds else None) or (ac_keywords[0] if ac_keywords else None) or title or "Unknown keyword"
    candidate = title or primary
    long_tail = unique_keep_order(ac_keywords + title_seeds)
    risk_flags = unique_keep_order(split_keywords(pick(row, "forbidden_flags")) + risk_flags_from_text(title, pick(row, "category_path"), pick(row, "detail_params"), pick(row, "sub_category")))
    source_ref = {"source": "SellerSprite" if "sellersprite" in str(source_file).lower() else "input_file", "file": source_file.name, "row": row_number, "collected_at": str(date.today()), "fields": list(row.keys())}
    history = extract_history(row, str(primary).strip(), source_file.name)
    return {
        "candidate": str(candidate).strip(),
        "primary_keyword": str(primary).strip(),
        "long_tail_keywords": long_tail,
        "title_seed_keywords": title_seeds,
        "verified_trend_keywords": sorted({p.get("keyword") for p in history if p.get("keyword")}),
        "source_site": source_site,
        "target_site": target_site,
        "identifiers": {"asin": pick(row, "asin"), "brand": brand, "product_url": pick(row, "product_url"), "image_url": pick(row, "image_url")},
        "category": {"path": pick(row, "category_path"), "main": pick(row, "main_category"), "sub": pick(row, "sub_category")},
        "market": {"monthly_sales": as_number(pick(row, "monthly_sales")), "monthly_sales_growth": as_number(pick(row, "monthly_sales_growth")), "monthly_revenue": as_number(pick(row, "monthly_revenue")), "search_volume": as_number(pick(row, "search_volume")), "new_product_count": as_number(pick(row, "new_product_count")), "new_product_sales_share": as_number(pick(row, "new_product_sales_share"))},
        "competition": {"top3_product_concentration": as_number(pick(row, "top3_product_concentration")), "top3_brand_concentration": as_number(pick(row, "top3_brand_concentration")), "top3_seller_concentration": as_number(pick(row, "top3_seller_concentration")), "click_concentration": as_number(pick(row, "click_concentration")), "conversion_concentration": as_number(pick(row, "conversion_concentration")), "spr": as_number(pick(row, "spr")), "cpc": as_number(pick(row, "cpc")), "review_count": as_number(pick(row, "review_count")), "rating": as_number(pick(row, "rating")), "seller_count": as_number(pick(row, "seller_count")), "variant_count": as_number(pick(row, "variant_count")), "amazon_owned_share": as_number(pick(row, "amazon_owned_share"))},
        "profit": {"price": as_number(pick(row, "price")), "cost": as_number(pick(row, "cost")), "fba_fee": as_number(pick(row, "fba_fee")), "gross_margin": as_number(pick(row, "gross_margin")), "return_rate": as_number(pick(row, "return_rate"))},
        "migration": {"au_fit_notes": "", "au_competitor_count": None, "seasonality_risk": "", "compliance_risk": pick(row, "compliance_risk") or ""},
        "risk": {"forbidden_flags": risk_flags, "patent_risk": pick(row, "patent_risk") or "", "assembly_complexity": "", "variant_complexity": "", "package_size_segment": pick(row, "package_size_segment")},
        "keyword_history": history,
        "source_refs": [source_ref],
    }


def flatten_json_records(data):
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ("candidates", "opportunities", "products", "keywords", "data"):
            value = data.get(key)
            if isinstance(value, list):
                return value
        return [data]
    return []


def read_records(path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix == ".xlsx":
        return read_xlsx(path)
    if suffix == ".json":
        return flatten_json_records(read_json(path))
    return []


def main():
    parser = argparse.ArgumentParser(description="Normalize AU product selection inputs")
    parser.add_argument("--input", required=True, help="CSV/XLSX/JSON file or folder")
    parser.add_argument("--topic", default="opportunities")
    parser.add_argument("--source-site", default="US")
    parser.add_argument("--target-site", default="AU")
    parser.add_argument("--output", help="Output run directory")
    args = parser.parse_args()
    input_path = Path(args.input)
    run_name = f"{args.topic}_{args.source_site}_to_{args.target_site}_{date.today().strftime('%Y%m%d')}".replace(" ", "_")
    output = Path(args.output) if args.output else Path("reports/au-product-selection") / run_name
    raw_dir = output / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    allowed = (".csv", ".json", ".xlsx")
    paths = [p for p in (input_path.rglob("*") if input_path.is_dir() else [input_path]) if p.suffix.lower() in allowed]
    candidates = []
    for source in paths:
        shutil.copy2(source, raw_dir / source.name)
        rows = read_records(source)
        for idx, row in enumerate(rows, start=1):
            if isinstance(row, dict):
                candidates.append(normalize_row(row, source, idx, args.source_site, args.target_site))
    data = {"metadata": {"topic": args.topic, "source_site": args.source_site, "target_site": args.target_site, "created_at": str(date.today()), "input_mode": "product_discovery" if any(p.suffix.lower() == ".xlsx" for p in paths) else "mixed"}, "candidates": candidates}
    (output / "data.json").write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    evidence = {"sources": [{"file": p.name, "path": str(raw_dir / p.name), "type": p.suffix.lower().lstrip(".")} for p in paths]}
    (output / "evidence.json").write_text(json.dumps(evidence, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Normalized {len(candidates)} candidates -> {output / 'data.json'}")


if __name__ == "__main__":
    main()
